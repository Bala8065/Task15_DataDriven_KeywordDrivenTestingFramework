import os
import openpyxl
from datetime import datetime

# Get absolute path dynamically
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(BASE_DIR, "data", "login_data.xlsx")

def read_test_data():
    """Read test data from Excel and return as list of dictionaries"""
    workbook = openpyxl.load_workbook(EXCEL_PATH)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        test_id, username, password, date, time, tester, result = row
        data.append({
            "test_id": test_id,
            "username": username,
            "password": password,
            "tester": tester
        })
    return data

def write_result(test_id, status):
    """Write test result, date and time back to Excel"""
    workbook = openpyxl.load_workbook(EXCEL_PATH)
    sheet = workbook.active
    from datetime import datetime
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row, 1).value == test_id:
            sheet.cell(row, 4).value = datetime.now().strftime("%Y-%m-%d")
            sheet.cell(row, 5).value = datetime.now().strftime("%H:%M:%S")
            sheet.cell(row, 7).value = status
            break
    workbook.save(EXCEL_PATH)
